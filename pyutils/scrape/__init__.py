""""""
import os
from typing import Dict
import pydoc
import math

import openpyxl
import openpyxl.utils.dataframe
import pandas as pd
import poppler
from poppler.cpp import image as poppler_image

from PIL import (
    Image,
)

def xlst_to_dataframes(filepath):
    wb = openpyxl.load_workbook(filepath)
    return {sn: pd.DataFrame(wb[sn].values) for sn in wb.sheetnames}


def dataframes_to_xlst(dfs: Dict[str, pd.DataFrame], filepath):
    wb = openpyxl.Workbook()
    wb.active
    for (nm, df) in dfs.items():
        ws = wb.create_sheet(nm)
        ws.title
        for row in openpyxl.utils.dataframe.dataframe_to_rows(
            df, index=True, header=True
        ):
            ws.append(row)
    wb.save(filepath)


def pdf_to_text(filename, out=None, page=False):
    doc = poppler.load_from_file(filename)
    text = ''
    for pg_idx in range(doc.pages):
        pg = doc.create_page(pg_idx)
        pg_text = pg.text()
        assert pg_text[-1] == '\x0c'
        text += pg_text[:-1]
    lines = [line.strip() for line in text.split('\n')]
    parsed_text = '\n'.join(lines)
    if page:
        pydoc.pager(parsed_text)
    if out is not None:
        with open(out, 'w') as f:
            f.write(parsed_text)
    return parsed_text


def _argb_to_rbga_bytes(argb_bytes: bytes):
    def transpose(some_bytes, ia, ib):
        byte_a = some_bytes[ia]
        some_bytes[ia] = some_bytes[ib]
        some_bytes[ib] = byte_a
        return some_bytes

    assert len(argb_bytes) % 4 == 0
    bytes_out = list(argb_bytes)
    for seg_idx in range(int(len(argb_bytes) / 4)):
        bytes_out = transpose(
            bytes_out,
            seg_idx*4,
            seg_idx*4+3,
        )
    return bytes(bytes_out)


def pdf_4_in_1(filename, out):

    if not os.path.exists(out):
        os.mkdir(out)
    if not os.path.isdir(out):
        raise Exception()
    else:
        fnames = os.listdir(out)
        for fname in fnames:
            if fname.startswith('page_') and fname.endswith('.tiff'):
                os.unlink(os.path.join(out, fname))

    doc = poppler.load_from_file(filename)
    doc_page_count = doc.pages
    for first_pg_idx in range(doc_page_count)[::4]:
        last_pg_idx = min(doc.pages-1, first_pg_idx+3)
        pg_idxs = list(range(first_pg_idx, last_pg_idx+1))
        pgs = [doc.create_page(pg_idx)
               for pg_idx in pg_idxs]
        rendr = poppler.PageRenderer()
        imgs = [rendr.render_page(pg) for pg in pgs]
        #
        img_formats = set([img.format for img in imgs])
        assert img_formats == {
            poppler_image.format_enum.argb32
        }
        img_dims = set([(img.width, img.height) for img in imgs])
        assert len(img_dims) == 1
        img_dims = list(img_dims)[0]
        #
        # Image.frombuffer
        pil_imgs = [
            Image.frombytes('RGBA',
                            img_dims,
                            _argb_to_rbga_bytes(img.data),
                            )
            for img in imgs]

        cimg = Image.new('RGBA', (img_dims[0]*2, img_dims[1]*2))
        ###
        # 4-tuple for boxes in PIL:
        #        (left, top, right, bottom)
        ##
        for idx, pil_img in enumerate(pil_imgs):
            if idx == 0:
                box = (0, 0, img_dims[0], img_dims[1])
            elif idx == 1:
                box = (img_dims[0], 0, img_dims[0]*2, img_dims[1])
            elif idx == 2:
                box = (0, img_dims[1], img_dims[0], img_dims[1]*2)
            elif idx == 3:
                box = (img_dims[0], img_dims[1], img_dims[0]*2, img_dims[1]*2)
            else:
                raise Exception()
            cimg.paste(pil_img, box=box)

        outfile_idx = str(int(first_pg_idx / 4)).zfill(
            len(str(int(math.ceil(doc_page_count / 4)))))
        outfile_name = f'page_{outfile_idx}.tiff'
        cimg.save(os.path.join(
            out, outfile_name))


