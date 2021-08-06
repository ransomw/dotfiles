"""-*-python-*-
suggested use:
place a dotted symlink to this
file in your home directory
  ln -s /.../pythonrc ~/.pythonrc
and set the environment variable
PYTHONSTARTUP
to the symlink
export PYTHONSTARTUP="${HOME}/.pythonrc"
see also:
https://docs.python.org/3/using/cmdline.html#envvar-PYTHONSTARTUP

### todo
* refactor (using rope) into pyutil
* conn gg Photos.
* droplet / bucket / gcloud
  |_--< low-Trust of thumb drives AND spinning
  ..---++intro to req., here <<<EO_ThHyerOrdTHOT
    prioritizing (FOT)
        connect generic stor (DO droplet)
           **after** _or_ **before**
        domain-specific stor (gg Photos)
EO_ThHyerOrdTHOT
* lint, gofmt-alike, trailing whitespace
  - pylintrc
  - pylint ASTeroid checker(s), if only for example
  - pre-commit
* auto-generate documentation
  - sphinx (?)
  - generate todos from this markdown format
* maintain and/or generate breakpoints list
  |_--> keep a bunch && toggle, don't waste
        time [staring at the computer, typ-
        -ing `breakpoint()` over && over.
  - also, cram a bunch of data into `tabulate`
    and inject a special variable into pdb.
* music player
  - "lock" mode: only responsive on specific keypress
    ( or sequence thereof) - displaying sequnce
    on other key.  thisisnot a security feature.
  - read metadata
  - access & cache album art, credits, etc.
    from wikipedia, allmusic, bandcamp, etc.
  - qtile integration (volume, disp toggle)
--- +eFmt(s)
  - annot playback (eg Tuple[timestamp...])
    (ie "verse|chorus|bridge", "measures
     _ to _") within an indiv piece
  - Dict[float, _]: playback speed to (eg) cpu usage
  - mix *at most* two (Th) << multiproc
    (eg "i can play [dylan|hendrix]'s _
     over _.  same song" 19',20' phenomenon)
    (ie global annotations)
  - equalizer & effects
* mic access
* camera access
* tab-complete python names (w/ rope)
  in at least one editor
* plivo (or twilio fallback) text msgs
* image view & editing
  - scaled .
* document view & editing
  - 4-in-1 pdf
  - pdf text extraction (non-OCR)
* generative music / ambient noise
  - binaural beats-type
  - bytebeat
  - etc.
*
......onecet an py-chemist has
     .... <<will i want more>>
/\//____/\/\//\\/\///\/\/\\/\//\/
../  |   /\  |\  |> get other phone
\/|\ |  /--\ | \ |droid (from storage)
/\| \| /    \|  \|> install android build
西 \ |                  on a laptop
   \_|__            (fuschia, 茹果渴能)

and leave this
  to   bury
"""
import typing
from typing import (
    Dict,
    List,
    Hashable,
    Tuple,
    Union,
    Sequence,
)
from collections import (
    namedtuple,
)
import datetime as dt
from datetime import (
    datetime as dto,
    timedelta,
)
from functools import (
    reduce,
    partial,
    wraps,
)
from inspect import (
    getmembers as gm,
    getsource,
    getsourcefile as gsf,
    getmodule,
    ismodule,
    isclass,
)
import operator as op
from operator import (
    add,
    itemgetter as ig,
    attrgetter,
)
import os
import sys
import os.path as pth
import random
from pprint import (
    # pp, 2.8+
    pprint,
)
import math
from copy import (
    deepcopy,
)

import pydoc
from pydoc import (
    pager,
)

import curses
import curses.textpad


import subprocess
from subprocess import (
    Popen,
    PIPE,
)

import shlex

import configparser

import pathlib

from io import (
    BytesIO,
    StringIO,
)

from warnings import warn

import zipfile

import gzip
from gzip import (
    GzipFile,
)

from tempfile import (
    gettempdir,
    mkdtemp,
)

import json

from importlib import (
    reload,
)

import shutil
from shutil import (
    copyfileobj,
)

import threading
import asyncio
import queue

import signal

import time

import ast

import socket
import ipaddress

import rlcompleter
import readline

### END stdlib imports import stdlib

readline.parse_and_bind("tab: complete")

###


def ppp(obj):
    sio = StringIO()
    pprint(obj, stream=sio)
    pager(sio.getvalue())


def gmn(*args, **kwargs):
    return [m[0] for m in gm(*args, **kwargs)]


def gs(*args, **kwargs):
    pydoc.pager(getsource(*args, **kwargs))


def getsourcefiles(*args, **kwargs):
    getsource_results = [getsource(arg, **kwargs)
                         for arg in args]
    sources_string = (
        '#<<<\n'+'\n#<<<\n'.join(getsource_results)
    )
    pydoc.pager(
        sources_string
    )


inc = lambda x: x + 1
dec = lambda x: x - 1


class Leaf:
    pass


class Branch:
    pass


class Tree:
    pass


class NameSpace:
    def __init__(self, obj):
        self._obj = obj

    def names(self):
        if ismodule(self._obj):
            return dir(self._obj)
        elif isclass(self._obj):
            return dir(self._obj)
        assert False, repr(self._obj) + " is not a module or class"

    def attrs(self):
        return map(partial(getattr, self._obj), self.names())

    def namespaces(self):
        [Namespace(attr) for attr in self.attrs() if any(juxt(ismodule, isclass)(attr))]


def pysearch_name(name, maxdepth=3):
    res = []
    permissive_getattr = excepts(
        (ModuleNotFoundError, AttributeError), partial(getattr), lambda _: None
    )

    def name_match(mname):
        return name in mname

    res += [sys.modules[mname] for mname in sys.modules.keys() if name_match(mname)]

    def search_class(cls):
        for mname in dir(cls):
            if name_match(mname):
                res.append(permissive_getattr(cls, mname))

    def search_module(module, depth):
        if depth > maxdepth:
            return
        if name in dir(module):
            res.append(permissive_getattr(module, name))
        for (mname, member) in [
            (mname, permissive_getattr(module, mname)) for mname in dir(module)
        ]:
            if not member:
                continue
            if name_match(mname):
                res.append(member)
            if isinstance(member, type):
                search_class(member)
            if ismodule(member):
                search_module(member, depth + 1)

    for mname in list(sys.modules.keys()):
        search_module(sys.modules[mname], 0)
    return res


ls = os.listdir


def cat(filepath, retstr=False):
    with open(filepath) as f:
        fstr = f.read()
    if retstr:
        return fstr
    pydoc.pager(fstr)


run = partial(
    subprocess.run,
    stdout=subprocess.PIPE,
    stderr=subprocess.PIPE,
    # shell=True,
    # check=True,
)



def run_1(cmd: Union[str, Sequence], timeout=None, stream=sys.stdout, err_stream=sys.stderr, input_str=None,):

    _cmd_before_coerce = cmd #dbgref
    if isinstance(cmd, (str,)):
        warn("run_1 coerce untested")
        cmd = shlex.split(cmd)
    if not isinstance(cmd, (list,)):
        cmd = list(cmd)


    def input_writer(proc, input_lines):
        for input_line in input_lines.split('\n'):
            input_line_bytes = bytes(input_line, 'utf-8')
            proc.stdin.writelines(
                [input_line_bytes]
            )
        proc.stdin.close()

    def output_reader(proc):
        for line in iter((proc.stdout.readline if proc.stdout else lambda: b""), b""):
            print(line.decode("utf-8"), file=stream, end="")
        return

    def err_reader(proc):
        for line in iter((proc.stderr.readline if proc.stderr else lambda: b""), b""):
            print(line.decode("utf-8"), file=err_stream, end="")
        return
        # todo: copyfileobj with .decode("utf-8") wrapper
        proc_strm = proc.stderr or b""

    proc_kwargs = {}
    if input_str is not None:
        proc_kwargs['stdin'] = subprocess.PIPE
    proc = Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, **proc_kwargs)
    tout = threading.Thread(target=output_reader, args=(proc,))
    terr = threading.Thread(target=err_reader, args=(proc,))
    if input_str is not None:
        tin = threading.Thread(target=input_writer, args=(proc,input_str,))
        tin.start()
    tout.start()
    terr.start()
    proc.wait(timeout=timeout)
    tout.join()
    terr.join()
    if input_str is not None:
        tin.join()


config = configparser.ConfigParser()
config.read(
    [
        os.path.join(os.path.dirname(os.path.realpath(__file__)), "python.conf"),
    ]
)


class PipInstallException(Exception):
    pass


# todo: --cache-dir option,
#    possibly following pre-commit cache strategy
# todo(???): build wheels
def pip_install(package_name):
    name_to_specifier = {
        name: config["package-specifiers"][name]
        for name in config["package-specifiers"]
    }
    if package_name not in name_to_specifier:
        if os.getenv("PY_DBG_IMPORTS"):
            breakpoint()
        raise PipInstallException("unknown package", (package_name,))
    specifier = name_to_specifier[package_name]
    cmd = ["pip", "install",] + (['-e',] if specifier.startswith('git+https://') else []) + [specifier,]
    res = subprocess.run(cmd)
    if res.returncode == 0:
        return
    raise PipInstallException("install failed", (res,))


###
_VENV_DIR = pth.join(str(pathlib.Path().home()), ".pyvenv")
_DEFAULT_VENV = pth.join(_VENV_DIR, "default")
if sys.prefix == sys.base_prefix:
    if not os.getenv("PY_CREATE_VENV"):
        raise Exception("not in venv.  set PY_CREATE_VENV to create")
    venv.create(_DEFAULT_VENV)
    print(". " + pth.join(_DEFAULT_VENV, "bin", "activate"))
    exit()


class ImportBlocker(object):
    def __init__(self):
        self.module_names = set()
        self.package_names = set()

    def find_module(self, fullname, path=None):
        if fullname.split(".")[0] in self.package_names:
            return self
        if fullname in self.module_names:
            return self
        return None

    def exec_module(self, mdl):
        # return an empty namespace
        return {}

    def create_module(self, spec):
        return None


import_blocker = ImportBlocker()
sys.meta_path.append(import_blocker)

AUTO_DBG=False


def my_except_hook(exctype, value, traceback):
    if exctype is KeyboardInterrupt:
        print("see you later!")
    sys.__excepthook__(exctype, value, traceback)
    if AUTO_DBG:
        # breakpoint()
        import pdb
        pdb.pm() # post-mortem

def install_package(name):
    pass


sys.excepthook = my_except_hook

while True:
    try:
        import toolz
        import toolz.functoolz as ftlz
        import toolz.itertoolz as itlz
        import toolz.dicttoolz as dtlz
        from toolz.functoolz import (
            compose_left,
            excepts,
            compose,
            curry,
            flip,
            juxt,
            thread_last,
        )
        from toolz.itertoolz import (
            accumulate,
            concat,
            concatv,
            cons,
            diff,
            first,
            isdistinct,
            groupby,
            mapcat,
            nth,
            unique,
        )
        from toolz.dicttoolz import (
            keymap,
            valmap,
            itemmap,
        )
        from toolz.curried import (
            get,
            keyfilter,
            valfilter,
            itemfilter,
        )
        import numpy as np

        import pandas as pd
        import openpyxl
        import openpyxl.utils.dataframe
        import requests
        import npyscreen

        # without Qt install (see python.conf)
        # gui `pymol.lauch([])` is inoperative
        # and this package is at most useful
        # to parse file formats (protien database -- .pdb), etc.
        import pymol

        import urwid
        import tkinter

        import astor
        import rope
        import rope.base.project
        from rope.base import libutils as rope_libutils
        from rope.base.project import (
            Project as RopeProject,
        )
        from rope.base.resources import (
            File as RopeFile,
            Folder as RopeFolder,
            Resource as RopeResource,
        )
        from rope.base.pyobjectsdef import (
            PyModule as RopePyModule,
        )
        from rope.base.codeanalyze import (
            SourceLinesAdapter as RopeSourceLinesAdapter,
        )
        from rope.base.change import (
            ChangeSet as RopeChangeSet
        )
        import rope.refactor.move
        from rope.refactor.move import (
            MoveGlobal as RopeMoveGlobal,
            MoveModule as RopeMoveModule,
        )
        import rope.refactor.multiproject
        import rope.contrib.generate
        import astor

        from libqtile import qtile

        sys.path.append(os.path.dirname(os.path.realpath(__file__)))
        import pyutils
        from pyutils.file_browse import (simple_file_browser_urwid,)
        from pyutils.calendar import (ics_cal_busy_times_this_week,)
        import pyutils.pyjuke as juke
        from pyutils.pastebin import pastebin_app
        from pyutils.cartography.osm import (
            osm_to_shp,
            plot_osm_shp,
        )
        from pyutils.cartography.test_osm import (
            TEST_MAP as TEST_OSM_MAP
        )
        from pyutils.py_alarm_call.dashbd import (
            dashbd,
        )
        from pyutils import cookbook as cb
        sys.path.pop()

    except ModuleNotFoundError as err:
        package_name = err.name
        try:
            print("attempting to install " + package_name)
            pip_install(package_name)
        except PipInstallException as ex:
            if os.getenv("PY_DBG_IMPORTS"):
                breakpoint()
            import_blocker.package_names.add(package_name)
        continue
    break


# reset to orig
# sys.excepthook = sys.__excepthook__

uninstalled_packages = import_blocker.package_names.copy()
if uninstalled_packages:
    print("uninstalled packages")
    print(uninstalled_packages)

###

plot_osm = plot_open_street_map_xml = compose(plot_osm_shp, osm_to_shp)
plot_osm_demo = lambda: plot_osm(TEST_OSM_MAP)


def view_pdb(db_id: str):
    fn = view_pdb
    garlic_exec = fn.garlic_exec
    cache_dir = fn.cache_dir_path
    if not pth.exists(cache_dir):
        os.mkdir(cache_dir)
    download_url_base = fn.download_url_base
    pdb_file_path = pth.join(
        cache_dir,
        "{db_id}.pdb".format(
            db_id=db_id,
        ),
    )
    pdb_archive_path = pth.join(
        cache_dir,
        "{db_id}.pdb.gz".format(
            db_id=db_id,
        ),
    )

    #    breakpoint()

    pdb_file_url = requests.compat.urljoin(
        download_url_base,
        "{db_id}.pdb".format(
            db_id=db_id,
        ),
    )
    pdb_archive_url = requests.compat.urljoin(
        download_url_base,
        "{db_id}.pdb.gz".format(
            db_id=db_id,
        ),
    )

    if not pth.exists(pdb_file_path) and not pth.exists(pdb_archive_path):
        res: request.models.Response = requests.get(pdb_archive_url, stream=True)
        assert res.status_code == 200
        chunks = res.iter_content(
            # read data in "whatever size the chunks are recv'd"
            # todo: extract personal dsl fn, documenting the "whatever"
            chunk_size=None,
        )
        with open(pdb_archive_path, "wb") as f:
            for chunk in chunks:
                f.write(chunk)

    if not pth.exists(pdb_file_path):
        with gzip.open(pdb_archive_path) as f_gz, open(pdb_file_path, "wb") as f:
            shutil.copyfileobj(f_gz, f)

    #    breakpoint()
    proc = Popen([garlic_exec, pdb_file_path])
    # todo: workaround garlic numeric keypad requirement
    #   - forward keys from python terminal
    #   - wrap window (X11 or other layer) to access keys
    #          while view selected.
    try:
        print("interrupt Ctl-c to exit")
        while True:
            time.sleep(1024)
    except KeyboardInterrupt:
        pass
    finally:
        proc.terminate()
        print("waiting on garlic exit")
        proc.wait()

    return

    # attempts at optimization
    with open() as f:
        with GzipFile(
            f,
            # default 'rb'.  no text mode opt here.  gzip.open() only
            mode="rb",
        ) as gf:
            gf.read1()


view_pdb.download_url_base = "https://files.rcsb.org/download/"
view_pdb.garlic_exec = "garlic"
view_pdb.cache_dir_path = pth.join(gettempdir(), "view_pdb_cache_dir")


# todo: scrape wikipedia page to pdb code
#    https://en.wikipedia.org/wiki/Muscarinic_acetylcholine_receptor_M4
# todo: search wikipedia page (NLP+force(?))
#    to receptors (pages w/ pdb codes)


def xlst_to_dataframes(filepath):
    wb = openpyxl.load_workbook(filepath)
    return {sn: pd.DataFrame(wb[sn].values) for sn in wb.sheetnames}


def dataframes_to_xlst(dfs: Dict[str, pd.DataFrame], filepath):
    wb = openpyxl.Workbook()
    wb.active
    for (nm, df) in dfs.items():
        ws = wb.create_sheet(nm)
        ws.title
        for row in openpyxl.utils.dataframe.dataframe_to_rows(
            df, index=True, header=True
        ):
            ws.append(row)
    wb.save(filepath)

##>>>##

###
# simple time-boxing and -estimation setup
#
# todo:
#* persistent dict (to disk)
#* calc time estimate sums
#

# activity to minutes
_TIME_ESTIMATE = {}


def _espeak_text(
        text,
        vol=85,
        disp=True,
        async_run=False,
):
    if disp:
        rainbow_print(text)
    cmd = ['espeak',
           '-v', 'f4',
           '-p', '135',
           '-a', str(vol),
           #
           # text,
           ]
    if async_run:
        warn("async `_espeak_text` untested")
        run_1(
            cmd,
            input_str=text,
        )
    else:
        subprocess.run(
            cmd,
            input=bytes(text, 'utf-8'),
        )


def _xmessage_text(
        text,
        disp=True,
):
    if disp:
        rainbow_print(text)
    proc = subprocess.Popen(
        ['xmessage',
         text,
         ],
    )
    return proc



def print_time_estimates():
    print(_TIME_ESTIMATE)
    # todo: calculate sums


def add_time_estimate_task(
        key_or_ks,
        minutes=None,
):
    if isinstance(key_or_ks, (str,)):
        key = key_or_ks
        act = key #activity
        # if act in _TIME_ESTIMATE:
        #     raise ValueError(
        #         "activity already exists",
        #         (act,))
        minutes = (input(("how long will '"
                          +act+
                          "' take? >[minutes]> ")) if minutes is None else float(minutes))
        _TIME_ESTIMATE[act] = float(minutes)
    elif isinstance(key_or_ks, (list,)):
        ks = key_or_ks
        raise ValueError(
            "nested activities unimplemented",
            (ks,))
    else:
        raise RuntimeError(
            "unknown activity specification",
            (key_or_ks, type(key_or_ks),))


def do_time_estimated_task(
        key_or_ks=None,
        min_warning=2.,
        mute=True,
):
    if key_or_ks == None:
        print("given options")
        print("-------------")
        print_time_estimates()
        print("----")
        for v in _TIME_ESTIMATE.values():
            if not isinstance(v, (int,)):
                raise NotImplementedError(
                    "nested activities unimpl")
        key = ''
        while not key or key not in _TIME_ESTIMATE:
            # todo: tab-completion
            key = input((
                "what task do you want"
                " to work on? )]} "))
            if key not in _TIME_ESTIMATE:
                print(("unknown activity '"+
                       key
                       +"'. "
                       "activities to choose from"
                       " are:"))
                print(', '.join(["'"+act+"'" for act in
                           _TIME_ESTIMATE.keys()]))
        key_or_ks = key
    #
    if isinstance(key_or_ks, (str,)):
        key = key_or_ks
        act = key
        est_min = _TIME_ESTIMATE[key]
        if est_min is None:
            print("no time estimate for '"+act+"'")
        elif not isinstance(est_min, (int, float,)):
            raise ValueError(
                "nested activities unimplemented",
                (est_min,))
        msg_begin = "fine, take "+str(est_min)+" minutes"
        rainbow_print(msg_begin)
        if not mute:
            _espeak_text(msg_begin)
        sec_wait = float(60.*(float(est_min)-float(min_warning)))
        PROG_BAR_STEPS = 10
        step_sec_wait = sec_wait/float(PROG_BAR_STEPS)
        for _ in range(PROG_BAR_STEPS):
            # print("waiting "+str(step_sec_wait), file=sys.stderr,)
            time.sleep(step_sec_wait)
            print('.', end='')
            sys.stdout.flush()
        print()
        msg_warn = ("you have "+str(min_warning)+" minutes left to finish "+
                    act)
        rainbow_print(msg_begin)
        if not mute:
            _espeak_text(msg_warn)
        warn_xm_proc = _xmessage_text(msg_warn)
        sec_wait = 60.*min_warning
        step_sec_wait = sec_wait/float(PROG_BAR_STEPS)
        for _ in range(PROG_BAR_STEPS):
            time.sleep(step_sec_wait)
            print('.', end='')
            sys.stdout.flush()
        print()
        msg_end = ("time's up!  your estimate of "+str(est_min)+" minutes "
                   "to completely complete "+act+" has elapsed. ")
        rainbow_print(msg_begin)
        if not mute:
            _espeak_text(msg_end)
        end_xm_proc = _xmessage_text(msg_end)
        warn_xm_proc.terminate()
        warn_xm_proc.wait()
        finished = None
        while finished is None:
            finished_txt = input("did you finish??? >[yes/no]>> ").strip()
            if finished_txt not in ['yes', 'no',]:
                print("please answer 'yes' or 'no'")
                continue
            finished = finished_txt == 'yes'
        end_xm_proc.terminate()
        end_xm_proc.wait()
        if finished:
            del _TIME_ESTIMATE[key]
            # todo: pos/neg-reinforcement
        else:
            _TIME_ESTIMATE[key] = None
            print("your time-estimate has been discarded.  "
                  # ^ irl stuff.
                  "to add a new estimate:")
            print("add_time_estimate_task('"+key+"')")
    elif isinstance(key_or_ks, (list,)):
        ks = key_or_ks
        raise ValueError(
            "nested activities unimplemented",
            (ks,))
    else:
        raise RuntimeError(
            "unknown activity specification",
            (key_or_ks, type(key_or_ks),))

##<<<##


def edit_text_terminal_curses(text):
    # startup
    stdscr = curses.initscr()
    curses.noecho()
    curses.cbreak()
    stdscr.keypad(True)
    ###

    height, width = stdscr.getmaxyx()
    editwin = curses.newwin(height - 10, width - 10, 1, 1)
    editwin.scrollok(True)
    curses.textpad.rectangle(stdscr, 0, 0, height - 9, width - 9)
    stdscr.refresh()

    box = curses.textpad.Textbox(editwin, insert_mode=True)
    for char in text:
        box.do_command(char)
    # Let the user edit until Ctrl-G is struck.
    box.edit()

    # Get resulting contents
    message = box.gather()

    # exit
    curses.nocbreak()
    stdscr.keypad(False)
    curses.echo()
    curses.endwin()

    return message



def rainbow_print(text):
    # Codes listed are from ECMA-48, Section 8.3.117, p. 61.
    # extracted from /usr/local/share/zsh/5.8/functions/Misc/colors
    ##
    # Attribute codes:
    _ATTRS = {
        name: '\033['+code+'m'
        for (code, name) in (
      ('00', 'none'),
      ('01', 'bold'),
      ('02', 'faint'),
        ('22', 'normal'),
      ('03', 'standout'),
               ('23', 'no-standout'),
      ('04', 'underline'),
          ('24', 'no-underline'),
      ('05', 'blink'),
        ('25', 'no-blink'),
      ('07', 'reverse'),
        ('27', 'no-reverse'),
      ('08', 'conceal'),
        ('28', 'no-conceal'),
                )}
    # Text color codes:
    _COLORS = {
        name: '\033['+code+'m'
        for (code, name) in (
      ('30', 'black'),
        ('40', 'bg-black'),
      ('31', 'red'),
        ('41', 'bg-red'),
      ('32', 'green'),
        ('42', 'bg-green'),
      ('33', 'yellow'),
        ('43', 'bg-yellow'),
      ('34', 'blue'),
        ('44', 'bg-blue'),
      ('35', 'magenta'),
        ('45', 'bg-magenta'),
      ('36', 'cyan'),
        ('46', 'bg-cyan'),
      ('37', 'white'),
        ('47', 'bg-white'),
    )}
    _COLOR_DEFAULTS = {
        name: '\033['+code+'m'
        for (code, name) in (
      ('39', 'default'),
        ('49', 'bg-default'),
                )}
    color_names = [color_name for color_name in
                   _COLORS.keys()
                   if not color_name.startswith('bg-')]
    for color_name, char in zip(color_names*math.ceil(len(text)/len(color_names)), text):
        print(_COLORS[color_name]+char, end='')
    print(_ATTRS['none'])



# todo:
#    * kill line
#    * maintain AST in parallel with text
#        - arrays of lines, lines of arrays
#    * i18n input methods, mathsym input method
#    * undo
#    * see Edit.highlight re. copy and paste impl
#    * additional cursor motion commands
#      - search text
#      - goto line
#    * use urwid Command Map
#    * allow editing in webbrowser
#      (see calc.py example)
def edit_text_terminal_urwid(edit_text, second_edit_text=None,
    word_re=r'[A-z]'
):
    class EditBox(urwid.Filler):
        region_bdry_marker = '|'

        def __init__(self, *args, **kwargs):
            edit_text = kwargs.pop('edit_text', '')
            self.edit = urwid.Edit(
                edit_text=edit_text,
                multiline=True,
            )
            super(EditBox, self).__init__(self.edit, *args, **kwargs)
            self.region_bdry = None

        def keypress(self, size, key):
            clipboard = edit_text_terminal_urwid.clipboard
            if key in ["ctrl n",]:
                key = 'down'
            elif key in ["ctrl p",]:
                key = 'up'
            elif key in ["ctrl b",]:
                key = 'left'
            elif key in ["ctrl f",]:
                key = 'right'
            elif key in ["ctrl a",]:
                key = 'home'
                if False:
                    edit_text_before = self.edit.get_edit_text()[:self.edit.edit_pos]
                    try:
                        newline_before = edit_text_before.rindex('\n')
                    except ValueError:
                        newline_before = -1
                    self.edit.set_edit_pos(newline_before+1)
                    return
            elif key in ["ctrl e",]:
                key = 'end'
                if False:
                    edit_text_after = self.edit.get_edit_text()[self.edit.edit_pos:]
                    try:
                        newline_after = edit_text_after.index('\n')
                    except ValueError:
                        newline_after = len(edit_text_after)
                    self.edit.set_edit_pos(self.edit.edit_pos+newline_after)
                    return
            ###
            elif key in ["meta <",]:
                self.edit.set_edit_pos(0)
            elif key in ["meta >",]:
                self.edit.set_edit_pos(len(self.edit.get_edit_text()))
            elif key in ["ctrl k",]:
                pass
            elif key in ["ctrl d",]:
                key = "delete"
            elif key in ["ctrl _",]:
                pass
            ###
            elif key in [
                    # "ctrl space"
                       "<0>",]:
                if self.region_bdry is None:
                    self.region_bdry = self.edit.edit_pos
                    self.edit.insert_text(
                        self.region_bdry_marker)
                    return
            elif key in ["meta w",]:
                if self.region_bdry is None:
                    return
                if self.edit.edit_pos <= self.region_bdry:
                    cp_text = self.edit.get_edit_text()[
                        self.edit.edit_pos:self.region_bdry]
                else:
                    cp_text = self.edit.get_edit_text()[
                        self.region_bdry+1:self.edit.edit_pos]
                clipboard.append(cp_text)
                self.edit.set_edit_text(
                    self.edit.get_edit_text()[:self.region_bdry]
                    +
                    self.edit.get_edit_text()[self.region_bdry+1:])
                self.region_bdry = None
                return
            elif key in ["ctrl w",]:
                if self.region_bdry is None:
                    return
                if self.edit.edit_pos <= self.region_bdry:
                    cut_text = self.edit.get_edit_text()[
                        self.edit.edit_pos:self.region_bdry]
                    self.edit.set_edit_text(
                        self.edit.get_edit_text()[:self.edit.edit_pos]
                        +
                        self.edit.get_edit_text()[self.region_bdry+1:])
                else:
                    cut_text = self.edit.get_edit_text()[
                        self.region_bdry+1:self.edit.edit_pos]
                    self.edit.set_edit_text(
                        self.edit.get_edit_text()[:self.region_bdry]
                        +
                        self.edit.get_edit_text()[self.edit.edit_pos:])
                    self.edit.set_edit_pos(self.region_bdry)
                clipboard.append(cut_text)
                self.region_bdry = None
                return
            elif key in ["ctrl y",]:
                if clipboard:
                    self.edit.insert_text(itlz.last(
                        clipboard))
                return
            elif key in ["ctrl u",]:
                try:
                    clipboard.pop()
                except IndexError:
                    pass
                return
            return super(EditBox, self).keypress(size, key)


    edit_box = EditBox(
        "top",
        edit_text=edit_text,
    )
    if second_edit_text is not None:
        second_edit_box = EditBox(
            "top",
            edit_text=second_edit_text,
        )
        main_loop_box = urwid.Columns([
            urwid.LineBox(
                edit_box,
            ),
            urwid.LineBox(
                second_edit_box,
            ),
        ])
    else:
        second_edit_box = None
        main_loop_box = urwid.LineBox(
            edit_box,
        )
    ###
    def unhandled_keypress(k):
        #
        if k in ["f10",]:
            raise urwid.ExitMainLoop()
        elif k in ["f9",]:
            edit_text_terminal_urwid.exit_flags.append("save")
            raise urwid.ExitMainLoop()
        ###
    #
    loop = urwid.MainLoop(
        main_loop_box,
        unhandled_input=unhandled_keypress,
    )
    loop.run()
    exit_flags = edit_text_terminal_urwid.exit_flags
    edit_text_terminal_urwid.clipboard = []
    edit_text_terminal_urwid.exit_flags = []
    return {
        'text': edit_box.edit.get_edit_text(),
        'second_text': (None if second_edit_box is None
                        else second_edit_box.edit.get_edit_text()),
        'exit_flags': exit_flags,
        'save': 'save' in exit_flags,
    }
edit_text_terminal_urwid.clipboard = []
edit_text_terminal_urwid.exit_flags = []


def edit_file_terminal(filename):
    filename = os.path.realpath(filename)
    with open(filename) as f:
        text_before = f.read()
    res = edit_text_terminal_urwid(text_before)
    text_after = res['text']
    with open(filename, 'w') as f:
        f.write(text_after)


sfbrow_ur = simple_file_browser_urwid


def python_viewer_urwid(src):
    """ stepping-stone towoard src editor:
    use AST in parallel with text.
    - syntax highlighting
    - folding
    - goto definition
    - find occurences
    - opt line no.s
    """
    pass


def mouse_only_ui(cmd_arg):
    cmd = (shlex.split(cmd_arg)
            if isinstance(cmd_arg, (str,))
            else cmd_arg)
    if not isinstance(cmd, (list,)):
        raise ValueError()
    #
    app_state = {}
    def stdio_reader(file_obj_name):
        if file_obj_name not in ['stdout', 'stderr']:
            raise ValueError()
        proc = app_state['proc']
        file_obj = getattr(proc, file_obj_name)
        for line in iter((file_obj.readline
                          if file_obj else lambda: b""), b""):
            app_state[file_obj_name].append(
                line.decode("utf-8"))

    def on_go():
        if 'proc' in app_state and app_state['proc'].poll() is None:
            return
        proc = Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        app_state['proc'] = proc
        # workaround "pipe buffer deadlock"
        # also consider
        # * .communicate() with a future (?)
        # * passing a file object to stdout and stderr
        app_state['stdout'] = []
        app_state['stderr'] = []
        app_state['reader_threads'] = [
            threading.Thread(
                target=stdio_reader,
                args=('stdout',),
            ),
            threading.Thread(
                target=stdio_reader,
                args=('stderr',),
            ),
        ]
        for t in app_state['reader_threads']:
            t.start()

    def on_stop():
        if 'proc' not in app_state:
            return
        proc = app_state['proc']
        #
        for sig in [None, signal.SIGINT,
                signal.SIGKILL,
                signal.SIGTERM,]:
            if sig is not None:
                proc.send_signal(sig)
            if sig is None:
                if proc.poll() is None:
                    continue
                else:
                    break
            try:
                if sig is not None:
                    proc.wait(timeout=3)
            except subprocess.TimeoutExpired as ex:
                continue
            break
        else:
            raise RuntimeError("process didn't exit")
        #
        for t in app_state['reader_threads']:
            t.join()
        if proc.poll() != 0:
            print("error exit")
            print("stdout")
            print(('\n'.join(app_state['stdout'])
                   if app_state['stdout']
                   else "<None>"))
            print("stderr")
            print(('\n'.join(app_state['stderr'])
                   if app_state['stderr']
                   else "<None>"))
        del app_state['proc']
        del app_state['reader_threads']
        del app_state['stdout']
        del app_state['stderr']
    ###
    # ui
    root = tkinter.Tk()
    left_frame = tkinter.Frame(
            root,
#            relief="raised",
#            color="#200",
            )
    left_frame["bg"] = "purple"
    right_frame = tkinter.Frame(
            root,
            )
    right_frame["bg"] = "pink"
    left_frame.pack(
            side="left",
            fill="both",
            anchor="center",
            expand=True,
            )
    right_frame.pack(
            side="right",
            fill="y",
            )
    ##
    # https://stackoverflow.com/questions/42579927/rounded-button-tkinter-python
    # suggests using canvas.
    # elsewhere suggested to use image/bitmap
    class RoundButton(tkinter.Button, tkinter.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            # self.create_line(0, 0, 10, 10)
            self.create_line(*self.bbox("all"))
        #
        def _configure(self, *args, **kwargs):
            rv = super()._configure(*args, **kwargs)
            breakpoint()
            return rv
    ##
    go_btn = tkinter.Button(
            left_frame,
            text="Go!",
            command=on_go,
            relief="raised",
            )
    go_btn["bg"] = "#082"
    stop_btn = tkinter.Button(
            right_frame,
            text="stop",
            command=on_stop,
            )
    stop_btn["bg"] = "#086"
    go_btn.pack(
            side="bottom",
            # expand=True,
            fill="both",
            )
    stop_btn.pack(
            side="top",
            )
    root.mainloop()



def rope_move_fn_from_pythonrc(fn_name, pyutils_pkg_name):

    if '.' in pyutils_pkg_name:
        raise ValueError()
    pyutils_rootdir = pth.dirname(pyutils.__file__)
    pythonrc_rootdir = pth.dirname(pyutils_rootdir)
    # pyutils_project = RopeProject(pyutils_rootdir)
    pythonrc_project = RopeProject(pythonrc_rootdir)
    #
    pyutils_pkg_specifier = 'pyutils.'+pyutils_pkg_name

    def _close():
        pythonrc_project.close()

    def _ensure_pyutils_pkg(rope_project) -> RopeFolder:
        found_pkg: RopeFolder = rope_project.find_module(pyutils_pkg_specifier)
        if found_pkg is None:
            new_pkg: RopeFolder = rope.contrib.generate.create_package(
                rope_project, pyutils_pkg_specifier)
            pyutils_pkg = new_pkg
        else:
            pyutils_pkg = found_pkg
        assert isinstance(pyutils_pkg, (RopeResource,))
        assert pyutils_pkg.name == pyutils_pkg_name
        assert pyutils_pkg.is_folder()
        name_to_pkg_file = {f.name: f for f in pyutils_pkg.get_files()}
        assert '__init__.py' in name_to_pkg_file
        init_py: RopeFile = name_to_pkg_file['__init__.py']
        return pyutils_pkg


    def _str_to_ast(fstr, fname):
        """ astor.code_to_ast.parse_file extract """
        fstr = fstr.replace('\r\n', '\n').replace('\r', '\n')
        if not fstr.endswith('\n'):
            fstr += '\n'
        return ast.parse(fstr, filename=fname)

    def _rope_file_to_ast(rope_file: RopeFile) -> ast.Module:
        return _str_to_ast(rope_file.read(), rope_file.name)


    def _rope_file_fn_bounds(rope_file: RopeFile, func_name: str) -> ast.Module:
        file_str = rope_file.read()
        module_ast = _str_to_ast(file_str, rope_file.name)
        name_to_function_def: Dict[str, ast.FunctionDef] = {
            node.name: node for node in module_ast.body
            if isinstance(node, (ast.FunctionDef,))
        }
        if func_name not in name_to_function_def:
            raise ValueError()
        function_def = name_to_function_def[func_name]
        bounds = {attr_name: getattr(function_def, attr_name)
                  for attr_name in
                  ['lineno', 'end_lineno',
                   'col_offset', 'end_col_offset']}
        file_lines = file_str.split('\n')
        start_char = len('\n'.join(
            file_lines[:function_def.lineno-1]
        )) + function_def.col_offset + 1 # past trailing newline
        end_char = len('\n'.join(
            file_lines[:function_def.end_lineno-1]
        )) + function_def.end_col_offset + 1
        bounds.update(char=start_char, end_char=end_char)
        return bounds


    def _rope_module_fn_bounds(rope_module: RopePyModule, func_name: str) -> ast.Module:
        module_ast = rope_module.get_ast()
        name_to_function_def: Dict[str, ast.FunctionDef] = {
            node.name: node for node in module_ast.body
            if isinstance(node, (ast.FunctionDef,))
        }
        if func_name not in name_to_function_def:
            raise ValueError()
        function_def = name_to_function_def[func_name]
        bounds = {attr_name: getattr(function_def, attr_name)
                  for attr_name in
                  ['lineno', 'end_lineno',
                   'col_offset', 'end_col_offset']}
        module_lines_adapter: RopeSourceLinesAdapter = rope_module.lines
        file_lines = [module_lines_adapter.get_line(lineno)
                      for lineno in range(module_lines_adapter.length())]
        char = module_lines_adapter.get_line_start(
            function_def.lineno
        ) + function_def.col_offset
        end_char = module_lines_adapter.get_line_start(
            function_def.end_lineno
        ) + function_def.end_col_offset
        bounds.update(char=char, end_char=end_char)
        return bounds


    dest_pkg = _ensure_pyutils_pkg(pythonrc_project)


    pythonrc_module: RopePyModule = pythonrc_project.get_module('pythonrc')
    # pythonrc_resource: RopeFile = pythonrc_project.get_resource('pythonrc.py')
    pythonrc_resource: RopeFile = pythonrc_module.get_resource()

    #~ pythonrc_ast: ast.Module = _rope_file_to_ast(pythonrc_resource)
    # pythonrc_module.get_ast()

    def _make_changeset_for_one_fn(fn_name):
        # fn_bounds = _rope_file_fn_bounds(pythonrc_resource, fn_name)
        fn_bounds = _rope_module_fn_bounds(pythonrc_module, fn_name)
        if fn_bounds['col_offset'] != 0:
            Exception("expected a top-level function", (fn_name, fn_bounds,))
        move_offset = fn_bounds["char"] + len('def ')

        pythonrc_resource_str = pythonrc_resource.read()
        move_offset_from_index = (pythonrc_resource_str
                                  .index('\ndef '+fn_name) + len('\ndef '))

        if move_offset != move_offset_from_index:
            warn("ast and string index offsets differ")
            proceed = None
            while proceed is None:
                proceed_txt = input("continue anyway? >[yes/no]>> ").strip()
                if proceed_txt not in ['yes', 'no',]:
                    print("please answer 'yes' or 'no'")
                    continue
                proceed = proceed_txt == 'yes'
            if not proceed:
                _close()
                return



        # breakpoint()

        move_obj_for_typecheck = rope.refactor.move.create_move(
            pythonrc_project,
            pythonrc_module.get_resource(),
            move_offset,
        )
        if not isinstance(move_obj_for_typecheck, (RopeMoveGlobal,)):
            raise RuntimeError()

        class MoveGlobalKeepSrcImports(RopeMoveGlobal):

            def _source_module_changes(self, dest):
                placeholder = '__rope_moving_%s_' % self.old_name
                from rope.refactor.move import _ChangeMoveOccurrencesHandle
                handle = _ChangeMoveOccurrencesHandle(placeholder)
                from rope.refactor import occurrences
                occurrence_finder = occurrences.create_finder(
                    self.project, self.old_name, self.old_pyname)
                start, end = self._get_moving_region()
                from rope.refactor.move import ModuleSkipRenamer
                renamer = ModuleSkipRenamer(occurrence_finder, self.source,
                                            handle, start, end)
                source = renamer.get_changed_module()
                from rope.base import libutils
                pymodule = libutils.get_string_module(self.project, source, self.source)
                #~ source = self.import_tools.organize_imports(pymodule, sort=False)
                if handle.occurred:
                    pymodule = libutils.get_string_module(
                        self.project, source, self.source)
                    # Adding new import
                    from rope.refactor import importutils
                    source, imported = importutils.add_import(
                        self.project, pymodule, self._new_modname(dest), self.old_name)
                    source = source.replace(placeholder, imported)
                from rope.base.change import ChangeContents
                return ChangeContents(self.source, source)



        rope_move_obj = RopeMoveGlobal(
            pythonrc_project,
            pythonrc_module.get_resource(),
            move_offset,
        )
        rope_changes = rope_move_obj.get_changes(dest_pkg)
        rope_changes_description = rope_changes.get_description()

        my_move_obj = MoveGlobalKeepSrcImports(
            pythonrc_project,
            pythonrc_module.get_resource(),
            move_offset,
        )
        my_changes: RopeChangeSet = my_move_obj.get_changes(dest_pkg)
        my_changes_description = my_changes.get_description()

        x = my_changes
        xx = rope_changes
        y = my_changes_description
        yy = rope_changes_description

        return my_changes


    fn_changes = _make_changeset_for_one_fn(fn_name)
    fn_changes_description = fn_changes.get_description()


    validate_src_res = pythonrc_project.validate(pythonrc_module.get_resource())
    validate_dest_res = pythonrc_project.validate(dest_pkg)
    validate_project_res = pythonrc_project.validate(pythonrc_project.root)

    if (validate_src_res is not None or
        validate_dest_res is not None or
        validate_project_res is not None):
        raise Exception("validation fail")

    ##
    input('take a moment to view the changes')
    pydoc.pager(fn_changes_description)
    proceed = None
    while proceed is None:
        proceed_txt = input("perform the move? >[yes/no]>> ").strip()
        if proceed_txt not in ['yes', 'no',]:
            print("please answer 'yes' or 'no'")
            continue
        proceed = proceed_txt == 'yes'
    if not proceed:
        _close()
        return

    pythonrc_project.do(fn_changes)

    _close()
    return
    fn_ast: ast.FunctionDef = astor.code_to_ast(globals()[fn_name])
    pythonrc_ast: ast.Module = astor.code_to_ast.parse_file(
        pythonrc_resource.real_path)

    fns_changes = [_make_changeset_for_one_fn(fn_name)
                  for fn_name in fn_names]
    all_changes = reduce(lambda a, b: (a.add_change(b) or True) and a, fns_changes)



mv_fn = rope_move_fn_from_pythonrc


###

class NetmapError(Exception):
    pass


class NetmapParseError(ValueError, NetmapError):
    pass


def _get_own_ip():
    raise NotImplementedError()


def _is_port_open(
        ip_addr,
        port_no,
) -> bool:
    addr = (str(ip_addr), port_no,)
    try:
        conn = socket.create_connection(addr)
    except ConnectionRefusedError:
        return False
    conn.close()
    return True #...#
    sock = socket.socket(
        socket.AF_INET,
        socket.SOCK_STREAM | socket.SOCK_NONBLOCK)
    sock.connect(addr)


def _is_port_open_direct_lib(
        ip_addr,
        port_no,
) -> bool:
    addr = (str(ip_addr), port_no,)
    try:
        conn = socket.create_connection(addr)
    except ConnectionRefusedError:
        return False
    conn.close()
    return True #...#
    sock = socket.socket(
        socket.AF_INET,
        socket.SOCK_STREAM | socket.SOCK_NONBLOCK)
    sock.connect(addr)



def _port_scan_single_host(host, ports,
                           scan_method='high-lvl',):
    scan_method__create_connection = ['high-lvl', 'low-lvl',
                           'hi', 'lo',
                           'create_connection', 'direct_lib']
    scan_method__direct_lib = ['high-lvl', 'low-lvl',
                           'hi', 'lo',
                           'create_connection', 'direct_lib']
    assert scan_method in ['high-lvl', 'low-lvl',
                           'hi', 'lo',
                           'create_connection', 'direct_lib']
    rv = {}
    for port in ports:
        print("scanning port", port)
        single_port_res: bool = _is_port_open_direct_lib(host, port)
        single_port_res: bool = _is_port_open(host, port)

        rv.update({port: single_port_res})
    return rv


def port_scan(host_or_hosts, ports: List[int]):

    def _parse_host(maybe_host):
        try:
            host = str(ipaddress.ip_address(str(host_or_hosts)))
        except Exception as err:
            breakpoint() # get exception class
            raise NetmapParseError("", (err,))
        return host


    def _parse_hosts(hosts):
        if isinstance(hosts, (list, tuple, iter)):
            return [_parse_hosts(host)
                    for host in hosts]
        if isinstance(hosts, (str,)):
            pass


    def _coerce_ports(port_or_ports):
        ports = None
        if isinstance(port_or_ports, int):
            ports = [port_or_ports]
        if isinstance(port_or_ports, (list, tuple, set,)):
            ports = list(port_or_ports)
            if not all([isinstance(p, int) for p in ports]):
                raise NetmapParseError(
                    "recursive port coercion unimpl"
                )
        if ports is None:
            raise NetmapParseError()
        return ports


    single_host = None
    try:
        single_host = _parse_host(host_or_hosts)
    except NetmapParseError:
        pass
    if single_host is not None:
        hosts = [single_host]
    else:
        hosts = _parse_hosts(host_or_hosts)
    rv = {}
    for host in hosts:
        print("scanning host", host)
        single_host_res = _port_scan_single_host(host,
                                                 # ports,
                                                 _coerce_ports(ports),
                                                 )
        rv.update(**{str(host): single_host_res})

    breakpoint()

    return rv


def is_pastebin_up():
    port_scan(_get_own_ip(), [5003, 5005])


currfn = lambda: _is_port_open('192.168.1.64', 5005)
currfn = lambda: port_scan('192.168.1.64', 5005)
def _mount_unmounted_usb_thumb__freebsd():
    raise NotImplementedError()

###

import subprocess as sp
def _mount_unmounted_usb_thumb():
    if (sp.run('uname', stdout=sp.PIPE, check=True).stdout.decode('utf-8').strip()
        !=
        'FreeBSD'):
        raise NotImplementedError()
    return _mount_unmounted_usb_thumb__freebsd()

mount_usb = lambda: _mount_unmounted_usb_thumb


def _man_viewer__freebsd(name, sec=1):
    raise NotImplementedError()


def man_viewer():
    """
    interactive, follow links, ManManPages, ....
    """
    if (sp.run('uname', stdout=sp.PIPE, check=True).stdout.decode('utf-8').strip()
        !=
        'FreeBSD'):
        raise NotImplementedError()
    __man_viewer__freebsd()


#########################################


from time import (
    sleep as sync_sleep,
)
def make_it_BEEEp():
    def _beep_once():
        warn("beep unimpl")
        # terminal bell
    while True:
        _beep_once()
        sync_sleep(1)



currdefns = {defn.__name__: defn for defn
             in [
                 python_viewer_urwid,
                 make_it_BEEEp,
                 mount_usb,
                 ]}



for varname in [
        'mv_fn',
        ]:
    pass


# todo: use `del` to unclutter locals()


#########################################



def thread_loop_demo():


    mock_op = MagicMock()
    mock_op.side_effect = lambda _: random.random() > 0.5


    def thread_entry(thread_queue):
        pass


    thread_queue = queue.Queue()
    loop_thread = threading.Thread(target=thread_entry, args=(thread_queue,), daemon=True)
    loop_thread.start()
    for idx in range(10):
        thread_queue.put({"idx": idx})
